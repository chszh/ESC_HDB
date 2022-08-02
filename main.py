#!/usr/bin/python3
import os
import sys

def read_xslx(filepath,key):
#Takes in path to xslx file and key field
#key field is a field that is unique across all entries
#e.g building ID
#returns the header row as a list
#and the rest of the xslx data as a dictionary, built with key provided
#e.g 
# ID,postalcode,blkno
# 1,2,3
# 4,5,6
# -->
# {1:{postalcode:2,blkno:3},4:{postalcode:5,blkno:6}}
    import openpyxl
    ##Hides warning
    import warnings
    warnings.simplefilter("ignore")
    ##
    wb = openpyxl.load_workbook(filename=filepath)
    all_sheets = wb.sheetnames
    header_row = [i.value for i in wb[all_sheets[0]][1]]
    sheets_dict={}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        primary_key = key
        for row in ws.iter_rows(min_row=2):
            row_dict={}
            for idx,i in enumerate(header_row):
                row_dict[header_row[idx]]=str(row[idx].value)
            sheets_dict[row[header_row.index(primary_key)].value]=row_dict
    return header_row,sheets_dict

def read_mappings(filepath):
#Reads the mappings.config file
#returns mappings dictionary that maps the xsl field name to gdb field name
    import os
    import sys
    with open(filepath,"r") as f:
        lines = f.readlines()
        llines = [line.strip().split(",") for line in lines]
        mappings={}
        xsl_idx = llines[0].index('xslx')
        gdb_idx = llines[0].index('gdb')
        for line in llines[1:]:
            mappings[line[xsl_idx]] = line[gdb_idx]
    return mappings

def read_gdb(filepath,key,header_row,mappings={}):
#header row as an input defines what attributes to extract from model
#this is in case the model contains unneeded attributes, we don't want to extract them
#throws error if cannot open gdb
#returns dictionary of all features, built with key provided
    import sys
    from osgeo import ogr
    ogr.UseExceptions()
    driver = ogr.GetDriverByName("OpenFileGDB")
    gdb_path = filepath
    try:
        gdb = driver.Open(gdb_path, 0)
    except Exception as e:
        print(e)
        sys.exit()
    featsClassList = []
    for featsClass_idx in range(gdb.GetLayerCount()):
        featsClass = gdb.GetLayerByIndex(featsClass_idx)
        featsClassList.append(featsClass)
    features_dict={}
    primary_key = key
    for layer in featsClassList:
        for feature in layer:
            dic = feature.items()
            if 'NUM_TYPE' in dic:
                if dic['NUM_TYPE'] != "STRATA":
                    continue
            feature_dict={}
            for idx,i in enumerate(header_row):
                #for field in header row, extract from dic and place into feat_dict
                if i in dic:
                    feature_dict[i] = dic[i]
                elif i in mappings:
                    if mappings[i] in dic:
                    #in case field name not in model
                    #check if mapped field name in model
                        feature_dict[i] = dic[mappings[i]]
                else:
                    print(f"Field {i} does not exist for feature {dic}")
                    continue
            if primary_key in dic:
                features_dict[dic[primary_key]] = feature_dict
            elif mappings[primary_key] in dic:
                features_dict[dic[mappings[primary_key]]] = feature_dict
    del gdb
    return features_dict

def compare_row(a,b,getkey=False):
#Takes in dictionaries a and b
#They represent a row of data
#e.g a = {"name":"john","age":5} b = {"name":"john","age":6}
#if a and b have equal contents, returns 0
#if they are different, return 1
#if different and getkey=True, returns the field, followed by first differing contents as tuple
#e.g ("age",5,6)
    if set(a.keys()) != set(b.keys()):
        raise Exception("xslx and gdb have differing headers")
    for key in a.keys():
        aa=a[key].strip()
        bb=b[key].strip()
        if aa.isdigit() and bb.isdigit():
            aa=int(aa)
            bb=int(bb)
        if aa != bb:
            if getkey:
                return (key,aa,bb)
            return 1
    return 0

def compare_data(xslx_dict,gdb_dict):
#performs compare_row on all rows of the xslx dict and gdb dict
#returns list of compare results in the order of the xslx
    to_return = []
    for unit in xslx_dict.keys():
        if unit in gdb_dict:
            a=xslx_dict[unit]
            b=gdb_dict[unit]
            r = compare_row(a,b)
            if r==0:
                to_return.append((0,unit))
            elif r==1:
                key,a,b = compare_row(a,b,getkey=True)
                to_return.append((1,unit,key,a,b))
        else:
            to_return.append((2,unit))
    return to_return


def start_analysis(gdbpath,xslxpath,key,mappings_path = os.path.dirname(os.path.realpath(sys.argv[0]))+"/mappings.config"):
#reads gdb and xslx from Entry widgets and compares them
#possible results
#identical
#(0,{key value})
#found matching key entries but some field differs
#(1,{key value},{differing value A},{differing value B},{entry in xslx},{entry gdb})
#could not find xslx entry in gdb
#(2,{key value})
#returns as list of results (which are tuples)
    gdb_filepath = gdbpath
    xslx_filepath = xslxpath
    header_row,xslx_dict = read_xslx(xslx_filepath,key)
    mappings = read_mappings(mappings_path)
    gdb_dict = read_gdb(gdb_filepath,key,header_row,mappings=mappings)
    to_return = compare_data(xslx_dict,gdb_dict)
    return to_return

def download_report(reportEntry,gdbEntry,xslxEntry,keyEntry):
#runs analysis function that generates list of results
#writes results into report.txt in given directory
    import datetime
    from pathlib import Path
    epoch_string = str(datetime.datetime.now().timestamp())[:-2]
    report_filename=f"report_{epoch_string}.txt"
    report_filepath = Path(reportEntry.get()).joinpath(report_filename)
    gdbpath = gdbEntry.get()
    xslxpath = xslxEntry.get()
    key = keyEntry.get()
    analysis_result = start_analysis(gdbpath,xslxpath,key)
    to_write=[]
    for i in analysis_result:
        if i[0] == 0:
            to_write.append(f"FOUND_IN_GDB,{key},{i[1]}")
        elif i[0] == 1:
            to_write.append(f"MISMATCH,{key},{i[1]},\"{i[2]}\" != \"{i[3]}\"")
            # to_write.append(f"XSLXENTRY:{i[4]}")
            # to_write.append(f"GDBENTRY:{i[5]}") 
        elif i[0] == 2:
            to_write.append(f"NOT_FOUND_IN_GDB,{key},{i[1]}")
    with open(report_filepath,'w') as f:
        f.write('\n'.join(to_write))

def setDirName(entry):
#tkinter GUI function
    import tkinter.filedialog
    loc = tkinter.filedialog.askdirectory(initialdir = ".")
    entry.delete(0,'end')
    entry.insert(0,loc)
    entry.xview_moveto(1)

def setFileName(entry):
#tkinter GUI function
    import tkinter.filedialog
    loc = tkinter.filedialog.askopenfilename(initialdir = ".")
    entry.delete(0,'end')
    entry.insert(0,loc)
    entry.xview_moveto(1)

def tkinit():
    import tkinter
    root=tkinter.Tk()
    baseFrame = tkinter.Frame(root)
    baseFrame.grid()
    gdbLabel = tkinter.Label(baseFrame, text="gdb file:")
    gdbLabel.grid(row=0, column=0)
    xslxLabel = tkinter.Label(baseFrame, text="xslx file:")
    xslxLabel.grid(row=1, column=0)
    reportLabel = tkinter.Label(baseFrame, text="report file:")
    reportLabel.grid(row=2, column=0)
    keyLabel = tkinter.Label(baseFrame, text="Key:")
    keyLabel.grid(row=3, column=0)
    keyEntry = tkinter.Entry(baseFrame)
    keyEntry.grid(row=3, column=1, columnspan=2)
    keyEntry.insert(tkinter.END,"NUM_UNIT_ID")
    gdbEntry = tkinter.Entry(baseFrame)
    gdbEntry.grid(row=0, column=1, columnspan=2)
    gdbEntry.insert(tkinter.END,"gdb file location")
    xslxEntry = tkinter.Entry(baseFrame)
    xslxEntry.grid(row=1, column=1, columnspan=2)
    xslxEntry.insert(tkinter.END,"xslx file location")
    gdbPickerButton = tkinter.Button(baseFrame, text="Browse", command=lambda: setDirName(gdbEntry))
    gdbPickerButton.grid(row=0, column=3)
    xslxPickerButton = tkinter.Button(baseFrame, text="Browse", command=lambda: setFileName(xslxEntry))
    xslxPickerButton.grid(row=1, column=3)
    reportPickerButton = tkinter.Button(baseFrame, text="Browse", command=lambda: setDirName(reportEntry))
    reportPickerButton.grid(row=2, column=3)
    reportEntry = tkinter.Entry(baseFrame)
    reportEntry.grid(row=2, column=1, columnspan=2)
    reportEntry.insert(tkinter.END,"report file location")
    downloadButton = tkinter.Button(baseFrame, text="Analyze + Download", command=lambda: download_report(reportEntry,gdbEntry,xslxEntry,keyEntry))
    downloadButton.grid(row=5, column=1)
    quitButton = tkinter.Button(baseFrame, text="Quit", command=root.destroy, bg='Red', fg='Black')
    quitButton.grid(row=5, column=3)
    return root

if __name__ == "__main__":
    try:
        root=tkinit()
        root.mainloop()
    except Exception as e:
        print(f"{type(e)}: {e}")
        print("Press enter to exit")
        raw_input()

# gdb_filepath = "/home/cheze/SchoolRelated/term5/software_element/hdb_project/TPY_Batch12.gdb"
# xslx_filepath = "/home/cheze/SchoolRelated/term5/software_element/hdb_project/Unit Attributes(PIDB).xlsx"
# report_filepath = "/home/cheze/SchoolRelated/term5/software_element/hdb_project"
# key="NUM_UNIT_ID"
# header_row=["NUM_UNIT_ID","NUM_BLDNG_GL","NUM_BLK","NME_STREET","NUM_POSTAL_CODE","NUM_LEVEL","NUM_UNIT"]
# gdbEntry.delete(0,'end')
# gdbEntry.insert(0,gdb_filepath)
# xslxEntry.delete(0,'end')
# xslxEntry.insert(0,xslx_filepath)
# reportEntry.delete(0,'end')
# reportEntry.insert(0,report_filepath)
# root.mainloop()