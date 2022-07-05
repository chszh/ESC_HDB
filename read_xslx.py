import openpyxl

wb = openpyxl.load_workbook(filename='TestGDBExcel.xlsx')
all_sheets = wb.sheetnames
ws = wb['Sheet1']
header_row = [i.value for i in ws[1]]
primary_key = "Unit"
sheet_dict={}
for row in ws.iter_rows(min_row=2):
	row_dict={}
	for idx,i in enumerate(header_row):
		if i != primary_key:
			row_dict[header_row[idx]]=row[idx].value
	sheet_dict[row[header_row.index(primary_key)].value]=row_dict
print(sheet_dict)